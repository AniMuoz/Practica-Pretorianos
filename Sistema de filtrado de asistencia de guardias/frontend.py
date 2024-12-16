import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import datetime
import os
from PIL import Image, ImageTk

# Aquí va el código de tu lógica existente (el que proporcionaste)

nombres =[]
ruts = []
record = []
event = []
fecha = []
asistencia = []
inasitencia = []
control_insasitencia = [[], [], [], [], []]  # Cinco listas vacías
data = [[], [], []]
ci = [[], [], []]

f = []
e = []

directorio_raiz = os.path.dirname(os.path.abspath(__file__))
ruta_imagen = os.path.join(directorio_raiz, 'images', 'LOGOFOOTER.PNG')

# Para simplificar, aquí solo se incluyen las funciones necesarias
# Asegúrate de incluir todas las funciones de tu código original

def busca_archivos(folder, data):
    count = 0
    zz = 0
    ff = 0

    for filename in os.listdir(folder):
        if filename.endswith(".xlsx"):  # Procesar solo archivos .xlsx
            filepath = os.path.join(folder, filename)
            excel = openpyxl.load_workbook(filepath)
            hoja = excel.active

            evento = [hoja.cell(row=1, column=1).value]
            fecha = [hoja.cell(row=3, column=1).value]

            nombres = []
            ruts = []
            asistencia = []

            for i in range(6, hoja.max_row + 1):
                nombre = hoja.cell(row=i, column=3).value
                rut = hoja.cell(row=i, column=4).value
                asiste = hoja.cell(row=i, column=2).value

                if not nombre or not rut:  # Si nombre o rut están vacíos, termina la lectura
                    break
                
                nombres.append(nombre)
                ruts.append(rut)
                asistencia.append(asiste)

                f.append(fecha[0])
                e.append(evento[0])

            print("Nombres: ", nombres)
            print("Ruts: ", ruts)
            print("Asistencia: ", asistencia)
            print("-" * 50)

            ci[0].append(nombres)
            ci[1].append(ruts)
            ci[2].append(asistencia)

            #print(len(ruts))
            #while ff < len(ruts):
            #    if zz != 0 and count == 0:
            #        ff = zz
            #    f.append(fecha[0])
            #    e.append(evento[0])
            #    ff += 1
            #    count += 1
            #zz = len(ruts)

            # Añadir datos a `data` asegurándose de que no se repitan nombres
            for i, nombre in enumerate(nombres):
                if nombre not in data[0]:  # Si el nombre no está ya en `data[0]`
                    data[0].append(nombre)
                    data[1].append(ruts[i])
                    record_value = 1 if asistencia[i] == 1 else 0
                    data[2].append(record_value)
                else:
                    # Si el nombre ya existe, actualiza el registro
                    index = data[0].index(nombre)
                    if asistencia[i] == 1:
                        data[2][index] += 1

            excel.close()
    print(len(f))
    #print(ci, len(ci[0]),len(ci[2]),len(ci[2][2]), len(ci[0][0]))
    # Procesar inasistencias
    k = 0
    for j in range(len(ci[2])):
        for i in range(len(ci[2][j])):
            #print(i,j)
            #print(len(ci[2][j]))
            #print(ci[2][j][i])
            #print(len(f))
            if ci[2][j][i] == None:  # Si el registro de asistencia es 0
                inasitencia.append("Inasistencia")
                control_insasitencia[0].append(ci[0][j][i])  # Nombre
                control_insasitencia[1].append(ci[1][j][i])  # RUT
                control_insasitencia[2].append("Inasistencia")
                control_insasitencia[3].append(e[k])  # Evento
                control_insasitencia[4].append(f[k])  # Fecha
            k += 1

    # Ordenar de mayor a menor
    indices_ordenados = sorted(range(len(data[2])), key=lambda i: data[2][i], reverse=True)
    data_ordenada = [[sublista[i] for i in indices_ordenados] for sublista in data]

    # Ordenar de menor a mayor
    indices_ordenadosinv = sorted(range(len(data[2])), key=lambda i: data[2][i])
    data_ordenadainv = [[sublista[i] for i in indices_ordenadosinv] for sublista in data]

    return data_ordenada, data_ordenadainv

def mas_asistencia(data_ordenada, dia):
    #Excel ordenado de mas asistentes a menos
    if os.path.exists(f"Record_guardias{dia}.xlsx"):
        guardias = openpyxl.load_workbook(f"Record_guardias{dia}.xlsx")
    else:
        guardias = openpyxl.Workbook()

    hoja = guardias.active

    hoja['A1'] = 'Nombre'   
    hoja['B1'] = 'Rut'
    hoja['C1'] = 'Cantidad de eventos asistidos'

    for i in range(0, len(data_ordenada[0])):
        hoja.cell(row = i + 2, column = 1, value = data_ordenada[0][i])
        hoja.cell(row = i + 2, column = 2, value = data_ordenada[1][i])
        hoja.cell(row = i + 2, column = 3, value = data_ordenada[2][i])

    guardias.save(f"Record_guardias{dia}.xlsx")

def menos_asistencia(data_ordenadainv, dia):
    #Excel ordenada de menos asistentes a mas asistentes
    if os.path.exists(f"Menos_asistencia_guardias_{dia}.xlsx"):
        guardias = openpyxl.load_workbook(f"Menos_asistencia_guardias_{dia}.xlsx")
    else:
        guardias = openpyxl.Workbook()

    hoja = guardias.active

    hoja['A1'] = 'Nombre'
    hoja['B1'] = 'Rut'
    hoja['C1'] = 'Cantidad de eventos asistidos'

    for i in range(0, len(data_ordenadainv[0])):
        hoja.cell(row = i + 2, column = 1, value = data_ordenadainv[0][i])
        hoja.cell(row = i + 2, column = 2, value = data_ordenadainv[1][i])
        hoja.cell(row = i + 2, column = 3, value = data_ordenadainv[2][i])

    guardias.save(f"Menos_asistencia_guardias_{dia}.xlsx")

def inasistencia(dia):
    #Excel ordenada de menos asistentes a mas asistentes
    if os.path.exists(f"Inasistencia_guardias_{dia}.xlsx"):
        guardias = openpyxl.load_workbook(f"Inasistencia_guardias_{dia}.xlsx")
    else:
        guardias = openpyxl.Workbook()

    hoja = guardias.active

    hoja['A1'] = 'Nombre'
    hoja['B1'] = 'Rut'
    hoja['C1'] = 'Evento'
    hoja['D1'] = 'Fecha'

    for i in range(0, len(control_insasitencia[0])):
        hoja.cell(row = i + 2, column = 1, value = control_insasitencia[0][i])
        hoja.cell(row = i + 2, column = 2, value = control_insasitencia[1][i])
        hoja.cell(row = i + 2, column = 3, value = control_insasitencia[3][i])
        hoja.cell(row = i + 2, column = 4, value = control_insasitencia[4][i])

    guardias.save(f"Inasistencia_guardias_{dia}.xlsx")

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Filtrado de Asistencia de Guardias")
        self.root.geometry("700x500")

        # Frame principal
        main_frame = tk.Frame(root)
        main_frame.pack(fill="both", expand=True)

        # Título (lado izquierdo)
        #title_label = tk.Label(main_frame, text="Sistema de asistencia de guardias", font=("Arial", 16, "bold"))
        #title_label.pack(side="left", anchor="nw", padx=10, pady=10)

        # Imagen (arriba a la derecha)
        #self.image = Image.open("../images/LOGOFOOTER.PNG")  # Reemplaza con la ruta de tu imagen
        self.image = Image.open(ruta_imagen)
        self.image = self.image.resize((100, 120), Image.Resampling.LANCZOS)
        self.photo = ImageTk.PhotoImage(self.image)

        image_label = tk.Label(main_frame, image=self.photo)
        image_label.place(relx=1.0, rely=0.0, anchor="ne", x=-10, y=10)  # Posiciona la imagen

        # Texto y entrada (parte principal)
        self.label = tk.Label(main_frame, text="Indique la ruta donde están los eventos:")
        self.label.pack(anchor="center", pady=10)

        self.folder_path = tk.StringVar()
        self.entry = tk.Entry(main_frame, textvariable=self.folder_path, width=50)
        self.entry.pack(anchor="center", pady=5)

        self.browse_button = tk.Button(main_frame, text="Buscar Carpeta", command=self.browse_folder)
        self.browse_button.pack(anchor="center", pady=5)

        self.process_button = tk.Button(main_frame, text="Procesar Eventos", command=self.process_events)
        self.process_button.pack(anchor="center", pady=5)

        self.result_label = tk.Label(main_frame, text="")
        self.result_label.pack(anchor="center", pady=5)

        self.option_var = tk.IntVar()
        self.option_label = tk.Label(main_frame, text="Seleccione una opción:")
        self.option_label.pack(anchor="center", pady=5)

        self.option1 = tk.Radiobutton(main_frame, text="1. Obtener record de guardias", variable=self.option_var, value=1)
        self.option1.pack(pady=5)

        self.option2 = tk.Radiobutton(main_frame, text="2. Obtener guardias que menos asistencia tiene", variable=self.option_var, value=2)
        self.option2.pack(pady=5)

        self.option3 = tk.Radiobutton(main_frame, text="3. Qué evento faltó cada guardia", variable=self.option_var, value=3)
        self.option3.pack(pady=5)

        self.execute_button = tk.Button(main_frame, text="Ejecutar Opción", command=self.execute_option)
        self.execute_button.pack(anchor="center", pady=10)

        footer_frame = tk.Frame(root, bg="gray")
        footer_frame.pack(side="bottom", fill="x")

        footer_label = tk.Label(footer_frame, text="Derechos Reservados © 2024", bg="gray", fg="white", font=("Arial", 10))
        footer_label.pack(anchor="center", pady=5)


    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        self.folder_path.set(folder_selected)

    def process_events(self):
        folder = self.folder_path.get()
        if not folder:
            messagebox.showerror("Error", "Por favor, seleccione una carpeta.")
            return
        # Aquí llamas a la función que procesa los archivos
        self.orden, self.ordeninv = busca_archivos(folder, data)
        messagebox.showinfo("Éxito", "Eventos procesados correctamente.")

    def execute_option(self):
        #dia = str(datetime.date.today().year) + str(datetime.date.today().month) + str(datetime.date.today().day)
        fecha = datetime.date.today()
        dia = str(fecha.year) + str(fecha.month) + str(fecha.day)
        print("Codigo de dia: ", dia)
        opcion = self.option_var.get()
        if opcion == 1:
            mas_asistencia(self.orden, dia)
            messagebox.showinfo("Éxito", "Se generó el archivo Excel con los datos de asistencia.")
        elif opcion == 2:
            menos_asistencia(self.ordeninv, dia)
            messagebox.showinfo("Éxito", "Se generó el archivo Excel con los datos de menos asistencia.")
        elif opcion == 3:
            inasistencia(dia)
            messagebox.showinfo("Éxito", "Se generó el archivo Excel con los datos de inasistencia.")
        else:
            messagebox.showerror("Error", "Por favor, seleccione una opción válida.")

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()

