import openpyxl
import datetime
import os.path as path
import os

#variables
nombres =[]
ruts = []
record = []
event = []
inasitencia = []
control_insasitencia = [[], [], [], []]  # Cuatro listas vacías
data = []


def busca_archivos(folder, data):
#ciclo para leer todos los archivos de la ruta
    for filename in os.listdir(folder):
        if filename.endswith(".xlsx"):  # Procesar solo archivos .xlsx
            filepath = os.path.join(folder, filename)
        
            excel = openpyxl.load_workbook(filepath) #abre excel
        
            hoja = excel.active #abre la hoja de excel
        
            evento = [hoja.cell(row=2,column=1).value] #se rescata el titulo del evento
            print(evento[0])

            nombres = [hoja.cell(row=i,column=3).value for i in range(5,hoja.max_row+1)] #se rescata el nombre de los gaurdias
            print("Nombres: ",nombres)

            ruts = [hoja.cell(row=i,column=4).value for i in range(5,hoja.max_row+1)] #se rescata ruts de los guardias
            print("Ruts: ",ruts)

            asistencia = [hoja.cell(row=i,column=2).value for i in range(5,hoja.max_row+1)] #se rescata asistencia al evento de los guardias
            print("Asistencia (1 = si | 2 = no): ", asistencia)
            
            print("-"*50)
            
            # Procesar asistencia
            for i in range(len(asistencia)):
                if asistencia[i] == None:  # Si la asistencia es 0
                    inasitencia.append("Inasistencia")  # Añadir a la lista de inasistencia
                    control_insasitencia[0].append(nombres[i])
                    control_insasitencia[1].append(ruts[i])
                    control_insasitencia[2].append("Inasistencia")
                    control_insasitencia[3].append(evento[0])  # Usar el título del evento
                    print(f"Inasistencia al evento: {nombres[i]}")  # Imprimir encontrado
            print("-"*50)
            # Añadir datos a data
            if len(data) == 0:
                data = [nombres, ruts, record]
            else:
                if len(data[0]) != len(nombres):  # Asegúrate de que data[0] y nombres tienen longitudes diferentes
                    for i in range(len(nombres)):
                        if nombres[i] not in data[0]:  # Solo añade si no está ya en data[0]
                            print("añadiendo:", nombres[i])
                            data[0].append(nombres[i])  # Añade el nombre
                            data[1].append(ruts[i])     # Añade el rut correspondiente
                            if record[i] == None:
                                record.insert(i, 0)
                            data[2].append(record[i])
                        
            excel.close()
            excel.save(filename)
        for i in range(0, len(data[0])):
            if len(record) == 0 :
                record.append(0)
            if asistencia[i] == 1:
                if record[i] == 0:
                    record.insert(i, 1)
                else:
                    record[i] = record[i] + 1
        
    if len(record) > len(data[0]):
        record.pop(-1)

    data[2].pop()

    #Print de debuging
    #print("Guardias con inasistencia:", inasitencia)
    #print(data[0])
    #print(data[2])
    #print(data[2])
    #print(data[3])
    #print("\nDatos insacistencia: ", control_insasitencia)

    #Ordenar de mayor a menor
    indices_ordenados = sorted(range(len(data[2])), key=lambda i: data[2][i], reverse=True)

    # Aplicamos el orden a cada sublista
    data_ordenada= [[sublista[i] for i in indices_ordenados] for sublista in data]

    #Ordenar de menor a mayor
    indices_ordenadosinv = sorted(range(len(data[2])), key=lambda i: data[2][i], reverse=False)

    # Aplicamos el orden a cada sublista
    data_ordenadainv= [[sublista[i] for i in indices_ordenadosinv] for sublista in data]

    return data_ordenada,data_ordenadainv



def mas_asistencia(data_ordenada, dia):
    #Excel ordenado de mas asistentes a menos
    if path.exists(f"Record_guardias{dia}.xlsx"):
        guardias = openpyxl.load_workbook(f"Record_guardias{dia}.xlsx")
    else:
        guardias = openpyxl.Workbook()

    hoja = guardias.active

    hoja['A1'] = 'Nombre'   
    hoja['B1'] = 'Rut'
    hoja['C1'] = 'Cantidad de eventos asistidos'

    for i in range(0, len(data_ordenada[0])):
        hoja.cell(row = i + 2, column = 1, value = data_ordenada[0][i])
        hoja.cell(row = i + 2, column = 2, value = data_ordenada[1][i])
        hoja.cell(row = i + 2, column = 3, value = data_ordenada[2][i])

    guardias.save(f"Record_guardias{dia}.xlsx")


def menos_asistencia(data_ordenadainv, dia):
#Excel ordenada de menos asistentes a mas asistentes
    if path.exists(f"Menos_asistencia_guardias_{dia}.xlsx"):
        guardias = openpyxl.load_workbook(f"Menos_asistencia_guardias_{dia}.xlsx")
    else:
        guardias = openpyxl.Workbook()

    hoja = guardias.active

    hoja['A1'] = 'Nombre'
    hoja['B1'] = 'Rut'
    hoja['C1'] = 'Cantidad de eventos asistidos'

    for i in range(0, len(data_ordenadainv[0])):
        hoja.cell(row = i + 2, column = 1, value = data_ordenadainv[0][i])
        hoja.cell(row = i + 2, column = 2, value = data_ordenadainv[1][i])
        hoja.cell(row = i + 2, column = 3, value = data_ordenadainv[2][i])

    guardias.save(f"Menos_asistencia_guardias_{dia}.xlsx")


def inasistencia(dia):
#Excel ordenada de menos asistentes a mas asistentes
    if path.exists(f"Inasistencia_guardias_{dia}.xlsx"):
        guardias = openpyxl.load_workbook(f"Inasistencia_guardias_{dia}.xlsx")
    else:
        guardias = openpyxl.Workbook()

    hoja = guardias.active

    hoja['A1'] = 'Nombre'
    hoja['B1'] = 'Rut'
    hoja['C1'] = 'Evento'

    for i in range(0, len(control_insasitencia[0])):
        hoja.cell(row = i + 2, column = 1, value = control_insasitencia[0][i])
        hoja.cell(row = i + 2, column = 2, value = control_insasitencia[1][i])
        hoja.cell(row = i + 2, column = 3, value = control_insasitencia[3][i])

    guardias.save(f"Inasistencia_guardias_{dia}.xlsx")

def main():
    fecha = datetime.date.today()
    dia = str(fecha.year) + str(fecha.month) + str(fecha.day)
    print("Codigo de dia: ", dia)
    #ruta de archivos
    #Mi ruta
    #folder = r"C:\Users\Anibal M\Desktop\Pretorianos\Practica-Pretorianos\Sistema de filtrado de asistencia de guardias\eventos"
    folder = input("Indique la ruta donde estan los eventos ==> ")
    orden, ordeninv = busca_archivos(folder, data)
    contador = 1
    while contador == 1:
        print("Que accion desea hacer:\n 1.- Obtener record de guardias\n 2.- Obtener guardias que menos aistencia tiene\n 3.- Que evento falto cada guardia\n 4.- Salir")
        opcion = int(input("Eliga el numero de la accion ==> "))
        while opcion >= 5 or opcion == 0:
            opcion = int(input("Seleccione una opcion valida ==> "))
        if opcion == 1:
            mas_asistencia(orden, dia)
            print("Se genero el archivo excel con los datos")
        if opcion == 2:
            menos_asistencia(ordeninv, dia)
            print("Se genero el archivo excel con los datos")
        if opcion == 3:
            inasistencia(dia)
            print("Se genero el archivo excel con los datos")
        if opcion == 4:
            print("Nos vemos!!!")
            break
        print("¿Desea realizar otra accion?")
        contador = int(input("1 = Si || 2 = No ==> "))
        while contador >= 3 or contador == 0:
            contador = int(input("Escoja una opcion valida\n1 = Si || 2 = No ==> "))
        if contador == 2:
            print("Nos vemos!!!")

main()