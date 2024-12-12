import os
import openpyxl
import datetime
import os.path as path
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter import Tk, ttk, filedialog, simpledialog
from PIL import Image, ImageTk  # Para manejar imágenes en la interfaz
from openpyxl.styles import Border, Side
from openpyxl.styles import NamedStyle

# Datos iniciales
data = [[], [], [], [], []]
monto_final = []

directorio_raiz = os.path.dirname(os.path.abspath(__file__))
ruta_imagen = os.path.join(directorio_raiz, 'images', 'LOGOFOOTER.PNG')

fecha = datetime.date.today()
dia = str(fecha.year) + str(fecha.month).zfill(2) + str(fecha.day).zfill(2)

topicos = ['CONSUMOS BASICOS', 'TELEFONO E INTERNET', 'GASTOS COMUNES', 'ARRIENDO DE OFICINA', 'COMBUSTIBLE', 'ESCRITORIO Y OFICINA', 
    'ESTACIONAMIENTO', 'ARTICULOS DE ASEO', 'GASTOS DE REPRESENTACIÓN', 'VESTUARIO Y CALZADO', 'PASAJES, PEAJES Y CORREOS',
    'MANTENIMIENTO, REPARACIÓN Y SEGURIDAD', 'EQUIPAMIENTO', 'ALIMENTACIÓN']

# Funciones principales
def guardar_data(data, dia):
    guardar = openpyxl.Workbook()
    hoja1 = guardar.active
    for i in range(len(data[1])):
        hoja1.cell(row=i + 1, column=1, value=data[0][i])
        hoja1.cell(row=i + 1, column=2, value=data[1][i])
        hoja1.cell(row=i + 1, column=3, value=data[2][i])
        hoja1.cell(row=i + 1, column=4, value=data[3][i])
        hoja1.cell(row=i + 1, column=5, value=data[4][i])
    filename = f"Datos_guardados_no_procesados_{dia}.xlsx"
    guardar.save(filename)
    messagebox.showinfo("Guardar datos", f"Datos guardados en {filename}")

def recuperar_data(data):
    filename = filedialog.askopenfilename(title="Seleccione el archivo de recuperación", filetypes=[("Archivos Excel", "*.xlsx")])
    if not filename:
        return

    try:
        excel = openpyxl.load_workbook(filename)
        hoja2 = excel.active
        for i in range(1, hoja2.max_row + 1):
            topi = hoja2.cell(row=i, column=1).value
            prove = hoja2.cell(row=i, column=2).value
            nbole = hoja2.cell(row=i, column=3).value
            febole = hoja2.cell(row=i, column=4).value
            mon = hoja2.cell(row=i, column=5).value

            if not topi or not prove or not nbole or not febole or not mon:
                break

            data[0].append(topi)
            data[1].append(prove)
            data[2].append(nbole)
            data[3].append(febole)
            data[4].append(int(mon))

        excel.close()
        messagebox.showinfo("Recuperar datos", "Datos recuperados con éxito")
        actualizar_tabla()
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo recuperar el archivo: {e}")

def procesar_datos(data, dia, topicos):
    guardias = openpyxl.Workbook()
    hoja = guardias.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    accounting_style = NamedStyle(name="accounting", number_format='$#')

    #hoja[f'E'].style = accounting_style

    hoja.column_dimensions['B'].width = 60
    hoja.column_dimensions['C'].width = 15
    hoja.column_dimensions['D'].width = 15
    hoja.column_dimensions['E'].width = 25

    mes = simpledialog.askstring("Procesar datos", "Ingrese el mes para procesar los datos:")
    if not mes:
        return

    hoja['B3'] = 'EMPRESA: PRETORIANOS SEGURIDAD'
    hoja['B4'] = 'DIRECCIÓN: MANUEL BULNES Nº 920, OFICINA 208, QUILPUÉ'
    hoja['B6'] = f'DETALLE GENERAL DE GASTOS MES {mes.upper()} DEL AÑO {fecha.year}'
    hoja['B6'].font = Font(bold=True, size=12)

    hoja['B8'] = 'CLASIFICACION DEL GASTO'
    hoja['B8'].font = Font(bold=True, size=12)
    hoja['B8'].border = thin_border
    hoja['C8'] = 'MONTO ($)'
    hoja['C8'].font = Font(bold=True, size=12)
    hoja['C8'].border = thin_border

    for i in range(9,24):
        hoja[f'B{i}'].border = thin_border

    hoja['B9'] = 'CONSUMOS BASICOS'		
    hoja['B10'] = 'TELEFONO E INTERNET'		
    hoja['B11'] = 'GASTOS COMUNES'		
    hoja['B12'] = 'ARRIENDO DE OFICINA'		
    hoja['B13'] = 'COMBUSTIBLE'		
    hoja['B14'] = 'ESCRITORIO Y OFICINA'		
    hoja['B15'] = 'ESTACIONAMIENTO'
    hoja['B16'] = 'ARTICULOS DE ASEO'		
    hoja['B17'] = 'GASTOS DE REPRESENTACIÓN'		
    hoja['B18'] = 'VESTUARIO Y CALZADO'		
    hoja['B19'] = 'PASAJES, PEAJES Y CORREOS'		
    hoja['B20'] = 'MANTENIMIENTO, REPARACIÓN Y SEGURIDAD'		
    hoja['B21'] = 'EQUIPAMIENTO'		
    hoja['B22'] = 'ALIMENTACIÓN'

    hoja['B23'] = 'TOTAL GASTOS DEL MES'
    hoja['B23'].font = Font(bold=True, size=12)
    hoja['C23'] = '=SUM(C9:C22)'
    hoja['C23'].font = Font(bold=True, size=12)
    hoja['C23'].style = accounting_style
    hoja['C23'].border = thin_border

    hoja['D26'] = 'FREDDY ANDRES MUÑOZ OLIVARES'
    hoja['D26'].font = Font(bold=True, size=12)
    hoja['D26'].alignment = Alignment(horizontal='center', vertical='center')
    hoja['D27'] = 'GERENTE GENERAL'
    hoja['D27'].font = Font(bold=True, size=12)
    hoja['D27'].alignment = Alignment(horizontal='center', vertical='center')

    fila = 30

    for i, topico in enumerate(topicos):
        hoja.cell(row=fila, column=2, value=f'DETALLE GASTOS EN {topico} MES {mes.upper()} DEL AÑO {fecha.year}').font = Font(bold=True, size=12)

        x_inicial = fila + 2
        hoja.cell(row = x_inicial, column = 2, value = "PROVEEDOR").font = Font(bold=True, size=11)
        hoja[f'B{x_inicial}'].alignment = Alignment(horizontal='center', vertical='center')
        hoja[f'B{x_inicial}'].border = thin_border
        hoja.cell(row = x_inicial, column = 3, value = "N° DE BOLETA").font = Font(bold=True, size=11)
        hoja[f'C{x_inicial}'].alignment = Alignment(horizontal='center', vertical='center')
        hoja[f'C{x_inicial}'].border = thin_border
        hoja.cell(row = x_inicial, column = 4, value = "FECHA").font = Font(bold=True, size=11)
        hoja[f'D{x_inicial}'].alignment = Alignment(horizontal='center', vertical='center')
        hoja[f'D{x_inicial}'].border = thin_border
        hoja.cell(row = x_inicial, column = 5, value = "MONTO ($)").font = Font(bold=True, size=11)
        hoja[f'E{x_inicial}'].alignment = Alignment(horizontal='center', vertical='center')
        hoja[f'E{x_inicial}'].border = thin_border

        contador = 1

        hoja.cell(row = x_inicial + contador, column = 2, value = "-").border = thin_border
        hoja.cell(row = x_inicial + contador, column = 3, value = "-").border = thin_border
        hoja.cell(row = x_inicial + contador, column = 4, value = "-").border = thin_border
        hoja.cell(row = x_inicial + contador, column = 5, value = 0).border = thin_border

        x_final = x_inicial
        for j in range(len(data[1])):
            if data[0][j] == topico:
                hoja.cell(row=x_inicial + contador, column=2, value=data[1][j]).border = thin_border
                hoja.cell(row=x_inicial + contador, column=3, value=data[2][j]).border = thin_border
                hoja.cell(row=x_inicial + contador, column=4, value=data[3][j]).border = thin_border
                hoja.cell(row=x_inicial + contador, column=5, value=data[4][j]).border = thin_border
                hoja[f'E{x_inicial + contador}'].style = accounting_style
                contador += 1

                x_final += 1

        hoja[f'B{x_final + 1}'].border = thin_border
        hoja[f'C{x_final + 1}'].border = thin_border
        hoja[f'D{x_final + 1}'].border = thin_border
        hoja[f'E{x_final + 1}'].border = thin_border

        hoja.cell(row = x_final + 2, column = 2, value = "VALOR TOTAL").font = Font(bold=True, size=12)
        hoja.merge_cells(f'B{x_final + 2}:D{x_final + 2}')
        hoja[f'B{x_final + 2}'].alignment = Alignment(horizontal='center', vertical='center')
        hoja[f'B{x_final + 2}'].border = thin_border
        hoja[f'C{x_final + 2}'].border = thin_border
        hoja[f'D{x_final + 2}'].border = thin_border
        hoja.cell(row = x_final + 2, column = 5, value = f'=IF(E{x_final}=0,0,SUM(E{x_inicial + 1}:E{x_final}))').font = Font(bold=True, size=12)
        hoja[f'E{x_final + 2}'].style = accounting_style
        hoja[f'E{x_final + 2}'].border = thin_border

        monto_final.append(f'E{x_final + 2}')

        fila += contador + 6

    for i in range (9, 23):
        hoja[f'C{i}'].style = accounting_style
        hoja.cell(row = i, column = 3, value = f'={monto_final[i - 9]}').border = thin_border
        

    filename = f"Detalle_gastos_pretorianos_{mes}_{dia}.xlsx"
    guardias.save(filename)
    messagebox.showinfo("Procesar datos", f"Datos procesados y guardados en {filename}")

def agregar_datos():
    topico = combo_topicos.get()
    proveedor = entry_proveedor.get()
    nboleta = entry_nboleta.get()
    fecha_boleta = entry_fecha_boleta.get()
    monto = entry_monto.get()

    if not (topico and proveedor and nboleta and fecha_boleta and monto):
        messagebox.showwarning("Advertencia", "Debe completar todos los campos")
        return

    try:
        monto = int(monto)
    except ValueError:
        messagebox.showerror("Error", "El monto debe ser un número entero")
        return

    data[0].append(topico)
    data[1].append(proveedor)
    data[2].append(nboleta)
    data[3].append(fecha_boleta)
    data[4].append(monto)

    actualizar_tabla()
    limpiar_campos()

def actualizar_tabla():
    for row in tree.get_children():
        tree.delete(row)

    for i in range(len(data[0])):
        tree.insert("", "end", values=(data[0][i], data[1][i], data[2][i], data[3][i], data[4][i]))

def limpiar_campos():
    combo_topicos.set("")
    entry_proveedor.delete(0, tk.END)
    entry_nboleta.delete(0, tk.END)
    entry_fecha_boleta.delete(0, tk.END)
    entry_monto.delete(0, tk.END)

# Configuración de la interfaz gráfica
root = tk.Tk()
root.title("Gestión de Gastos")

frame_logo = ttk.Frame(root)
frame_logo.pack(pady=10)
frame_logo.place(relx=1.0, rely=0.0, anchor="ne", x=-10, y=10)

try:
    img_logo = Image.open(ruta_imagen)  # Reemplaza 'logo.png' con la ruta de tu archivo de logo
    img_logo = img_logo.resize((100, 120))
    img_logo_tk = ImageTk.PhotoImage(img_logo)
    lbl_logo = ttk.Label(frame_logo, image=img_logo_tk)
    lbl_logo.image = img_logo_tk
    lbl_logo.pack()
except Exception as e:
    messagebox.showerror("Error", f"No se pudo cargar el logo: {e}")

frame_form = ttk.Frame(root)
frame_form.pack(pady=10)

lbl_topicos = ttk.Label(frame_form, text="Tópico:")
lbl_topicos.grid(row=0, column=0, padx=5, pady=5)
combo_topicos = ttk.Combobox(frame_form, values=topicos)
combo_topicos.grid(row=0, column=1, padx=5, pady=5)

lbl_proveedor = ttk.Label(frame_form, text="Proveedor:")
lbl_proveedor.grid(row=1, column=0, padx=5, pady=5)
entry_proveedor = ttk.Entry(frame_form)
entry_proveedor.grid(row=1, column=1, padx=5, pady=5)

lbl_nboleta = ttk.Label(frame_form, text="N° Boleta:")
lbl_nboleta.grid(row=2, column=0, padx=5, pady=5)
entry_nboleta = ttk.Entry(frame_form)
entry_nboleta.grid(row=2, column=1, padx=5, pady=5)

lbl_fecha_boleta = ttk.Label(frame_form, text="Fecha Boleta:")
lbl_fecha_boleta.grid(row=3, column=0, padx=5, pady=5)
entry_fecha_boleta = ttk.Entry(frame_form)
entry_fecha_boleta.grid(row=3, column=1, padx=5, pady=5)

lbl_monto = ttk.Label(frame_form, text="Monto:")
lbl_monto.grid(row=4, column=0, padx=5, pady=5)
entry_monto = ttk.Entry(frame_form)
entry_monto.grid(row=4, column=1, padx=5, pady=5)

btn_agregar = ttk.Button(frame_form, text="Agregar", command=agregar_datos)
btn_agregar.grid(row=5, column=0, columnspan=2, pady=10)

frame_table = ttk.Frame(root)
frame_table.pack(pady=10)

columns = ("Tópico", "Proveedor", "N° Boleta", "Fecha Boleta", "Monto")
tree = ttk.Treeview(frame_table, columns=columns, show="headings")
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=150)

tree.pack()

frame_buttons = ttk.Frame(root)
frame_buttons.pack(pady=10)

btn_guardar = ttk.Button(frame_buttons, text="Guardar Datos", command=lambda: guardar_data(data, dia))
btn_guardar.grid(row=0, column=0, padx=5)
btn_recuperar = ttk.Button(frame_buttons, text="Recuperar Datos", command=lambda: recuperar_data(data))
btn_recuperar.grid(row=0, column=1, padx=5)
btn_procesar = ttk.Button(frame_buttons, text="Procesar Datos", command=lambda: procesar_datos(data, dia, topicos))
btn_procesar.grid(row=0, column=2, padx=5)

root.mainloop()
