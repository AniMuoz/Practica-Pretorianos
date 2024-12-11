import os
import openpyxl
import datetime
import os.path as path
from openpyxl.styles import Font

data = [[],[],[],[],[]]

fecha = datetime.date.today()
dia = str(fecha.year) + str(fecha.month) + str(fecha.day)
print("Codigo de dia: ", dia)

def datos(topicos, data):
    contador = 1
    while contador == 1:
        for i in range (len(topicos)):
            print(i + 1, ".- ", topicos[i])
        top = int(input("Elija el topico al que desee ingresar boleta ==> "))
        while top <= 0 or top > len(topicos):
            top = int(input("Elija el topico al que desee ingresar boleta (valor valido) ==> "))
        proveedor = input("Ingrese proveedor: ")
        nboleta = input("Ingrese numero de boleta: ")
        fecha_boleta = input("Ingrese fecha de la boleta: ")
        monto = int(input("Monto de la boleta: "))

        data[0].append(topicos[top - 1])
        data[1].append(proveedor)
        data[2].append(nboleta)
        data[3].append(fecha_boleta)
        data[4].append(monto)
        contador = int(input("Añadir otro valor? 1 = si | 0 = no ==> "))
        while contador > 1 or contador < 0:
            contador = int(input("Ingrese opcion valida 1 = si | 0 = no ==> "))

    return(data)
    

#def gastos(dia):
#Excel ordenada de menos asistentes a mas asistentes
if path.exists(f"Detalle_gastos_pretorianos_seguridad_{dia}.xlsx"):
    guardias = openpyxl.load_workbook(f"Detalle_gastos_pretorianos_seguridad_{dia}.xlsx")
else:
    guardias = openpyxl.Workbook()

hoja = guardias.active

mes = input("Escriba el mes en el que quiere hacer la nomina: ")

hoja['B3'] = 'EMPRESA: PRETORIANOS SEGURIDAD'
hoja['B4'] = 'DIRECCIÓN: MANUEL BULNES Nº 920, OFICINA 208, QUILPUÉ'
hoja['B6'] = f'DETALLE GENERAL DE GASTOS MES {mes} DEL AÑO 2024'
hoja['B6'].font = Font(bold=True, size=12)

hoja['B8'] = 'CLASIFICACION DEL GASTO'
hoja['B8'].font = Font(bold=True, size=12)
hoja['C8'] = 'MONTO ($)'
hoja['C8'].font = Font(bold=True, size=12)

hoja['B9'] = 'CONSUMOS BASICOS'		
hoja['B10'] = 'TELEFONO E INTERNET'		
hoja['B11'] = 'GASTOS COMUNES'		
hoja['B12'] = 'ARRIENDO DE OFICINA'		
hoja['B13'] = 'COMBUSTIBLE'		
hoja['B14'] = 'ESCRITORIO Y OFICINA'		
hoja['B15'] = 'ESTACIONAMIENTO'
hoja['B16'] = 'ARTICULOS DE ASEO'		
hoja['B17'] = 'GASTOS DE REPRESENTACIÓN'		
hoja['B18'] = 'VESTUARIO Y CALZADO'		
hoja['B19'] = 'PASAJES, PEAJES Y CORREOS'		
hoja['B20'] = 'MANTENIMIENTO, REPARACIÓN Y SEGURIDAD'		
hoja['B21'] = 'EQUIPAMIENTO'		
hoja['B22'] = 'ALIMENTACIÓN'	
hoja['B23'] = 'TOTAL GASTOS DEL MES'	

topicos = ['CONSUMOS BASICOS', 'TELEFONO E INTERNET', 'GASTOS COMUNES', 'ARRIENDO DE OFICINA', 'COMBUSTIBLE', 'ESCRITORIO Y OFICINA', 
        'ESTACIONAMIENTO', 'ARTICULOS DE ASEO', 'GASTOS DE REPRESENTACIÓN', 'VESTUARIO Y CALZADO', 'PASAJES, PEAJES Y CORREOS',
        'MANTENIMIENTO, REPARACIÓN Y SEGURIDAD', 'EQUIPAMIENTO', 'ALIMENTACIÓN']

#for i in range (9, 23):
#    valor = int(input(f"Monto de {topicos[i - 9]}: "))
#    hoja.cell(row = i, column = 3, value = valor)

hoja['C23'] = '=SUM(C9:C22)'

hoja['D26'] = 'FREDDY ANDRES MUÑOZ OLIVARES'
hoja['D26'].font = Font(bold=True, size=12)
hoja['D27'] = 'GERENTE GENERAL'
hoja['D27'].font = Font(bold=True, size=12)

fila = 30

#hoja['B30'] = f'DETALLE GASTOS EN {topicos[0]} MES {mes} DEL AÑO {fecha.year}'
#hoja.cell(row = fila, column = 2, value = f'DETALLE GASTOS EN {topicos[0]} MES {mes} DEL AÑO {fecha.year}').font = Font(bold=True, size=12)


data = datos(topicos, data)

print("Len data[1]", len(data[1]))
print("Len topicos", len(topicos))

for i in range(len(topicos)):

    #hoja['B30'] = f'DETALLE GASTOS EN {topicos[0]} MES {mes} DEL AÑO {fecha.year}'
    hoja.cell(row = fila, column = 2, value = f'DETALLE GASTOS EN {topicos[i]} MES {mes} DEL AÑO {fecha.year}').font = Font(bold=True, size=12)


    x_inicial = fila + 2
    hoja.cell(row = x_inicial, column = 2, value = "PROVEEDOR")
    hoja.cell(row = x_inicial, column = 3, value = "N° DE BOLETA")
    hoja.cell(row = x_inicial, column = 4, value = "FECHA")
    hoja.cell(row = x_inicial, column = 5, value = "MONTO ($)")

    x = 1
    contador = 1
    x_final = x_inicial
    for j in range(len(data[1])):
        #print(f"DETALLE GASTOS EN {topicos[i]} MES {mes} DEL AÑO {fecha.year}")

        #proveedor = input("Ingrese proveedor: ")
        #nboleta = input("Ingrese numero de boleta: ")
        #fecha_boleta = input("Ingrese fecha de la boleta: ")
        #monto = int(input("Monto de la boleta: "))
        if data[0][j] == topicos[i]:
            hoja.cell(row = x_inicial + contador, column = 2, value = data[1][j])
            hoja.cell(row = x_inicial + contador, column = 3, value = data[2][j])
            hoja.cell(row = x_inicial + contador, column = 4, value = data[3][j])
            hoja.cell(row = x_inicial + contador, column = 5, value = data[4][j])
            print(f"se añadio data en {topicos[i]}")

            contador += 1
    
            x_final += 1

    hoja.cell(row = x_final + 1, column = 2, value = "VALOR TOTAL").font = Font(bold=True, size=12)
    hoja.cell(row = x_final + 1, column = 5, value = f'=SUM(E{x_inicial + 1}:E{x_final})')

    fila = x_final + 4

guardias.save(f"Detalle_gastos_pretorianos_seguridad_{dia}.xlsx")
