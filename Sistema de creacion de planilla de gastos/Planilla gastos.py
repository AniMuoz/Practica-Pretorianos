import os
import openpyxl
import datetime
import os.path as path
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from PIL import Image, ImageTk

data = [[],[],[],[],[]]
monto_final = []

fecha = datetime.date.today()
dia = str(fecha.year) + str(fecha.month) + str(fecha.day)
print("Codigo de dia: ", dia)

topicos = ['CONSUMOS BASICOS', 'TELEFONO E INTERNET', 'GASTOS COMUNES', 'ARRIENDO DE OFICINA', 'COMBUSTIBLE', 'ESCRITORIO Y OFICINA', 
    'ESTACIONAMIENTO', 'ARTICULOS DE ASEO', 'GASTOS DE REPRESENTACIÓN', 'VESTUARIO Y CALZADO', 'PASAJES, PEAJES Y CORREOS',
    'MANTENIMIENTO, REPARACIÓN Y SEGURIDAD', 'EQUIPAMIENTO', 'ALIMENTACIÓN']

def datos(topicos, data):
    contador = 1
    while contador == 1:
        for i in range (len(topicos)):
            print(i + 1, ".- ", topicos[i])
        top = int(input("Elija el topico al que desee ingresar boleta ==> "))
        while top <= 0 or top > len(topicos):
            top = int(input("Elija el topico al que desee ingresar boleta (valor valido) ==> "))
        proveedor = input("Ingrese proveedor: ")
        nboleta = input("Ingrese numero de boleta: ")
        fecha_boleta = input("Ingrese fecha de la boleta: ")
        monto = int(input("Monto de la boleta: "))

        data[0].append(topicos[top - 1])
        data[1].append(proveedor)
        data[2].append(nboleta)
        data[3].append(fecha_boleta)
        data[4].append(monto)
        contador = int(input("Añadir otro valor? 1 = si | 0 = no ==> "))
        while contador > 1 or contador < 0:
            contador = int(input("Ingrese opcion valida 1 = si | 0 = no ==> "))

    return(data)
    

def guardar_data(data, dia):
    guardar = openpyxl.Workbook()
    hoja1 = guardar.active
    for i in range(len(data[1])):
        hoja1.cell(row = i + 1, column = 1, value = data[0][i])
        hoja1.cell(row = i + 1, column = 2, value = data[1][i])
        hoja1.cell(row = i + 1, column = 3, value = data[2][i])
        hoja1.cell(row = i + 1, column = 4, value = data[3][i])
        hoja1.cell(row = i + 1, column = 5, value = data[4][i])
    guardar.save(f"Datos_guardados_no_procesados_{dia}.xlsx")

def recuperar_data(data):
    recupera = input("Ingrese el nombre del archivo de recuperacion con su extencion ==> ")
    excel = openpyxl.load_workbook(recupera)
    hoja2 = excel.active
    for i in range(1, hoja2.max_column + 1):
        print(i)
        topi = hoja2.cell(row = i, column = 1).value
        prove = hoja2.cell(row = i, column = 2).value
        nbole = hoja2.cell(row = i, column = 3).value
        febole = hoja2.cell(row = i, column = 4).value
        mon = hoja2.cell(row = i, column = 5).value

        print(topi, prove, nbole,febole, mon)

        if not topi or not prove or not nbole or not febole or not mon:
            break

        data[0].append(topi)
        data[1].append(prove)
        data[2].append(nbole)
        data[3].append(febole)
        data[4].append(int(mon))

    excel.save(recupera)
    return (data)

def gastos(dia, data, topicos):
    #if path.exists(f"Detalle_gastos_pretorianos_seguridad_{dia}.xlsx"):
    #    guardias = openpyxl.load_workbook(f"Detalle_gastos_pretorianos_seguridad_{dia}.xlsx")
    #else:
    #    guardias = openpyxl.Workbook()
    guardias = openpyxl.Workbook()

    hoja = guardias.active

    hoja.column_dimensions['B'].width = 60
    hoja.column_dimensions['C'].width = 15
    hoja.column_dimensions['D'].width = 15
    hoja.column_dimensions['E'].width = 25

    mes = input("Escriba el mes en el que quiere hacer la nomina: ")

    hoja['B3'] = 'EMPRESA: PRETORIANOS SEGURIDAD'
    hoja['B4'] = 'DIRECCIÓN: MANUEL BULNES Nº 920, OFICINA 208, QUILPUÉ'
    hoja['B6'] = f'DETALLE GENERAL DE GASTOS MES {mes} DEL AÑO 2024'
    hoja['B6'].font = Font(bold=True, size=12)

    hoja['B8'] = 'CLASIFICACION DEL GASTO'
    hoja['B8'].font = Font(bold=True, size=12)
    hoja['C8'] = 'MONTO ($)'
    hoja['C8'].font = Font(bold=True, size=12)

    hoja['B9'] = 'CONSUMOS BASICOS'		
    hoja['B10'] = 'TELEFONO E INTERNET'		
    hoja['B11'] = 'GASTOS COMUNES'		
    hoja['B12'] = 'ARRIENDO DE OFICINA'		
    hoja['B13'] = 'COMBUSTIBLE'		
    hoja['B14'] = 'ESCRITORIO Y OFICINA'		
    hoja['B15'] = 'ESTACIONAMIENTO'
    hoja['B16'] = 'ARTICULOS DE ASEO'		
    hoja['B17'] = 'GASTOS DE REPRESENTACIÓN'		
    hoja['B18'] = 'VESTUARIO Y CALZADO'		
    hoja['B19'] = 'PASAJES, PEAJES Y CORREOS'		
    hoja['B20'] = 'MANTENIMIENTO, REPARACIÓN Y SEGURIDAD'		
    hoja['B21'] = 'EQUIPAMIENTO'		
    hoja['B22'] = 'ALIMENTACIÓN'	
    hoja['B23'] = 'TOTAL GASTOS DEL MES'
    hoja['B23'].font = Font(bold=True, size=12)

    hoja['C23'] = '=SUM(C9:C22)'
    hoja['C23'].font = Font(bold=True, size=12)

    hoja['D26'] = 'FREDDY ANDRES MUÑOZ OLIVARES'
    hoja['D26'].font = Font(bold=True, size=12)
    hoja['D26'].alignment = Alignment(horizontal='center', vertical='center')
    hoja['D27'] = 'GERENTE GENERAL'
    hoja['D27'].font = Font(bold=True, size=12)
    hoja['D27'].alignment = Alignment(horizontal='center', vertical='center')

    fila = 30

    for i in range(len(topicos)):

        hoja.cell(row = fila, column = 2, value = f'DETALLE GASTOS EN {topicos[i]} MES {mes} DEL AÑO {fecha.year}').font = Font(bold=True, size=12)

        x_inicial = fila + 2
        hoja.cell(row = x_inicial, column = 2, value = "PROVEEDOR").font = Font(bold=True, size=11)
        hoja[f'B{x_inicial}'].alignment = Alignment(horizontal='center', vertical='center')
        hoja.cell(row = x_inicial, column = 3, value = "N° DE BOLETA").font = Font(bold=True, size=11)
        hoja[f'C{x_inicial}'].alignment = Alignment(horizontal='center', vertical='center')
        hoja.cell(row = x_inicial, column = 4, value = "FECHA").font = Font(bold=True, size=11)
        hoja[f'D{x_inicial}'].alignment = Alignment(horizontal='center', vertical='center')
        hoja.cell(row = x_inicial, column = 5, value = "MONTO ($)").font = Font(bold=True, size=11)
        hoja[f'E{x_inicial}'].alignment = Alignment(horizontal='center', vertical='center')

        x = 1
        contador = 1
        
        hoja.cell(row = x_inicial + contador, column = 2, value = "-")
        hoja.cell(row = x_inicial + contador, column = 3, value = "-")
        hoja.cell(row = x_inicial + contador, column = 4, value = "-")
        hoja.cell(row = x_inicial + contador, column = 5, value = 0)
        
        x_final = x_inicial
        for j in range(len(data[1])):
            if data[0][j] == topicos[i]:
                hoja.cell(row = x_inicial + contador, column = 2, value = data[1][j])
                hoja.cell(row = x_inicial + contador, column = 3, value = data[2][j])
                hoja.cell(row = x_inicial + contador, column = 4, value = data[3][j])
                hoja.cell(row = x_inicial + contador, column = 5, value = data[4][j])
                print(f"se añadio data en {topicos[i]}")

                contador += 1

                x_final += 1

        hoja.cell(row = x_final + 2, column = 2, value = "VALOR TOTAL").font = Font(bold=True, size=12)
        hoja.cell(row = x_final + 2, column = 5, value = f'=IF(E{x_final}=0,0,SUM(E{x_inicial + 1}:E{x_final}))').font = Font(bold=True, size=12)

        monto_final.append(f'E{x_final + 2}')

        fila = x_final + 5

    for i in range (9, 23):
        hoja.cell(row = i, column = 3, value = f'={monto_final[i - 9]}')

    guardias.save(f"Detalle_gastos_pretorianos_{mes}_{dia}.xlsx")
    #print(monto_final)

def main(data, dia, topicos):

    contador = 1

    while contador == 1:

        r = int(input("1.- Ingresar datos\n2.- Guardar datos sin procesar\n3.- Recuperar datos\n4.- Procesar datos\n==> "))

        while r > 4 or r < 1:
            r = int(input("Ingrese opcion valida del 1 al 4 ==> "))

        if r == 1:
            data = datos(topicos, data)
    
        if r == 2:
            guardar_data(data, dia)
    
        if r == 3:
            data = recuperar_data(data)

        if r == 4:
            gastos(dia ,data , topicos)

        contador = int(input("¿Desea seguir? 1 = Si | 0 = No ==> "))

        while contador > 1 or contador < 0:
            contador = int(input("Ingrese opcion valida 1 = si | 0 = no ==> "))

main(data, dia, topicos)