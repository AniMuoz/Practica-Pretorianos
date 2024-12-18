import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import datetime
import os

fecha = datetime.date.today()
dia = str(fecha.year) + str(fecha.month) + str(fecha.day)
print("Codigo de dia: ", dia)

data = []

def leer_base_datos(data):
    file = "Ejemplo_base_de_datos.xlsx"

    excel = openpyxl.load_workbook(file)
    hoja = excel.active

    nombres = []
    ruts = []

    for i in range(6, hoja.max_row + 1):
        nombre = hoja.cell(row=i, column=3).value
        rut = hoja.cell(row=i, column=4).value

        if not nombre or not rut:  # Si nombre o rut están vacíos, termina la lectura
            break
                
        nombres.append(nombre)
        ruts.append(rut)
    
    data.append(nombres)
    data.append(ruts)

    #print(nombres, ruts)

    excel.close()

    return(data)

def eleccion(data):
    guardias = [[],[]]

    cont = 1

    while cont == 1:

        for i in range(len(data[0])):
            print(i + 1, " .- ", data[0][i])

        elige = int(input("Elija el guardia que quiera añadir a la nomina: "))

        guardias[0].append(data[0][elige - 1])
        guardias[1].append(data[1][elige - 1])

        cont = int(input("Quiere añadir otro? 1 = Si | 0 = No ==> "))
        if cont < 0 or cont > 1:
            cont = int(input("Elija opcion valida 1 = Si | 0 = No ==> "))

    return guardias


def escribir_nomina(data):
    filename = input("Indique el nombre del evento: ")
    date = input("Indique fecha del evento: ")

    if os.path.exists(filename + ".xlsx"):
        nomina = openpyxl.load_workbook(filename + ".xlsx")
    else:
        nomina = openpyxl.Workbook()

    manejar = nomina.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    manejar.column_dimensions['A'].width = 5
    manejar.column_dimensions['B'].width = 5
    manejar.column_dimensions['C'].width = 35
    manejar.column_dimensions['D'].width = 20
    manejar.column_dimensions['E'].width = 15

    nom = "Nomina de guardias - Evento " + filename

    manejar['A1'] = nom
    manejar['A1'].font = Font(name = 'Arial Narrow', bold=True, size=12, underline='single')
    manejar['A1'].alignment = Alignment(horizontal='center', vertical='center')

    manejar['A3'] = date
    manejar['A3'].font = Font(name = 'Arial Narrow', bold=True, size=12, underline='single')

    manejar['A5'] = 'N°'
    manejar['A5'].font = Font(name = 'Arial Narrow', bold=True, size=12)
    manejar['A5'].border = thin_border
    manejar['B5'] = '0'
    manejar['B5'].font = Font(name = 'Arial Narrow', bold=True, size=12)
    manejar['B5'].border = thin_border
    manejar['C5'] = 'Apellidos y Nombres'
    manejar['C5'].font = Font(name = 'Arial Narrow', bold=True, size=12)
    manejar['C5'].border = thin_border
    manejar['D5'] = 'Ced. Idnt.'
    manejar['D5'].font = Font(name = 'Arial Narrow', bold=True, size=12)
    manejar['D5'].border = thin_border
    manejar['E5'] = 'Obs'
    manejar['E5'].font = Font(name = 'Arial Narrow', bold=True, size=12)
    manejar['E5'].border = thin_border

    manejar.merge_cells('A1:E1')

    for i in range(0, len(data[0])):
        manejar.cell(row = i + 6, column = 1, value = i + 1).border = thin_border
        manejar.cell(row = i + 6, column = 2).border = thin_border
        manejar.cell(row = i + 6, column = 3, value = data[0][i]).border = thin_border
        manejar.cell(row = i + 6, column = 4, value = data[1][i]).border = thin_border
        manejar.cell(row = i + 6, column = 5).border = thin_border 

    nomina.save(filename + ".xlsx")

data = leer_base_datos(data)

guardias = eleccion(data)

escribir_nomina(guardias)