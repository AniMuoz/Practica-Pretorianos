import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
import datetime
import os
import tkinter as tk
from tkinter import messagebox, filedialog

def leer_base_datos():
    file = filedialog.askopenfilename(title="Seleccione el archivo de base de datos", filetypes=[("Archivos Excel", "*.xlsx")])
    if not file:
        messagebox.showerror("Error", "Debe seleccionar un archivo válido.")
        return []

    excel = openpyxl.load_workbook(file)
    hoja = excel.active

    nombres = []
    ruts = []

    for i in range(6, hoja.max_row + 1):
        nombre = hoja.cell(row=i, column=3).value
        rut = hoja.cell(row=i, column=4).value

        if not nombre or not rut:
            break

        nombres.append(nombre)
        ruts.append(rut)

    excel.close()

    return [nombres, ruts]

def escribir_nomina(data, evento, fecha):
    filename = evento

    if os.path.exists(filename + ".xlsx"):
        nomina = openpyxl.load_workbook(filename + ".xlsx")
    else:
        nomina = openpyxl.Workbook()

    manejar = nomina.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    manejar.column_dimensions['A'].width = 5
    manejar.column_dimensions['B'].width = 5
    manejar.column_dimensions['C'].width = 35
    manejar.column_dimensions['D'].width = 20
    manejar.column_dimensions['E'].width = 15

    nom = "Nomina de guardias - Evento " + filename

    manejar['A1'] = nom
    manejar['A1'].font = Font(name='Arial Narrow', bold=True, size=12, underline='single')
    manejar['A1'].alignment = Alignment(horizontal='center', vertical='center')

    manejar['A3'] = fecha
    manejar['A3'].font = Font(name='Arial Narrow', bold=True, size=12, underline='single')

    manejar['A5'] = 'N°'
    manejar['A5'].font = Font(name='Arial Narrow', bold=True, size=12)
    manejar['A5'].border = thin_border
    manejar['B5'] = '0'
    manejar['B5'].font = Font(name='Arial Narrow', bold=True, size=12)
    manejar['B5'].border = thin_border
    manejar['C5'] = 'Apellidos y Nombres'
    manejar['C5'].font = Font(name='Arial Narrow', bold=True, size=12)
    manejar['C5'].border = thin_border
    manejar['D5'] = 'Ced. Idnt.'
    manejar['D5'].font = Font(name='Arial Narrow', bold=True, size=12)
    manejar['D5'].border = thin_border
    manejar['E5'] = 'Obs'
    manejar['E5'].font = Font(name='Arial Narrow', bold=True, size=12)
    manejar['E5'].border = thin_border

    manejar.merge_cells('A1:E1')

    for i in range(0, len(data[0])):
        manejar.cell(row=i + 6, column=1, value=i + 1).border = thin_border
        manejar.cell(row=i + 6, column=2).border = thin_border
        manejar.cell(row=i + 6, column=3, value=data[0][i]).border = thin_border
        manejar.cell(row=i + 6, column=4, value=data[1][i]).border = thin_border
        manejar.cell(row=i + 6, column=5).border = thin_border

    nomina.save(filename + ".xlsx")
    messagebox.showinfo("Éxito", f"La nómina se ha guardado en {filename}.xlsx")

def agregar_guardias(data, root):
    guardias = [[], []]

    def añadir_guardia():
        seleccionado = lista_guardias.curselection()
        if seleccionado:
            indice = seleccionado[0]
            guardias[0].append(data[0][indice])
            guardias[1].append(data[1][indice])
            lista_seleccionados.insert(tk.END, f"{len(guardias[0])}. {data[0][indice]}")

    def eliminar_guardia():
        seleccionado = lista_seleccionados.curselection()
        if seleccionado:
            indice = seleccionado[0]
            del guardias[0][indice]
            del guardias[1][indice]
            lista_seleccionados.delete(indice)
            actualizar_numeracion()

    def actualizar_numeracion():
        lista_seleccionados.delete(0, tk.END)
        for idx, nombre in enumerate(guardias[0]):
            lista_seleccionados.insert(tk.END, f"{idx + 1}. {nombre}")

    def finalizar():
        evento = entrada_evento.get()
        fecha = entrada_fecha.get()
        if not evento or not fecha:
            messagebox.showerror("Error", "Debe ingresar el nombre y la fecha del evento.")
            return
        escribir_nomina(guardias, evento, fecha)

    frame = tk.Frame(root)
    frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    tk.Label(frame, text="Guardias disponibles:", font=("Arial", 14)).grid(row=0, column=0, padx=5, pady=5)
    lista_guardias = tk.Listbox(frame, height=15, width=40, font=("Arial", 12))
    lista_guardias.grid(row=1, column=0, padx=5, pady=5)

    for idx, nombre in enumerate(data[0]):
        lista_guardias.insert(tk.END, f"{idx + 1}. {nombre}")

    boton_agregar = tk.Button(frame, text="Añadir", font=("Arial", 12), command=añadir_guardia)
    boton_agregar.grid(row=1, column=1, padx=5, pady=5)

    tk.Label(frame, text="Guardias seleccionados:", font=("Arial", 14)).grid(row=0, column=2, padx=5, pady=5)
    lista_seleccionados = tk.Listbox(frame, height=15, width=40, font=("Arial", 12))
    lista_seleccionados.grid(row=1, column=2, padx=5, pady=5)

    boton_eliminar = tk.Button(frame, text="Eliminar", font=("Arial", 12), command=eliminar_guardia)
    boton_eliminar.grid(row=2, column=2, padx=5, pady=5)

    tk.Label(frame, text="Nombre del evento:", font=("Arial", 12)).grid(row=3, column=0, padx=5, pady=5)
    entrada_evento = tk.Entry(frame, font=("Arial", 12))
    entrada_evento.grid(row=3, column=1, padx=5, pady=5)

    tk.Label(frame, text="Fecha del evento:", font=("Arial", 12)).grid(row=4, column=0, padx=5, pady=5)
    entrada_fecha = tk.Entry(frame, font=("Arial", 12))
    entrada_fecha.grid(row=4, column=1, padx=5, pady=5)

    boton_finalizar = tk.Button(frame, text="Finalizar", font=("Arial", 12), command=finalizar)
    boton_finalizar.grid(row=5, column=1, padx=5, pady=5)

def iniciar_aplicacion():
    root = tk.Tk()
    root.title("Gestión de Guardias")
    root.geometry("960x600")

    def cargar_datos():
        data = leer_base_datos()
        if data:
            agregar_guardias(data, root)

    boton_cargar = tk.Button(root, text="Cargar Base de Datos", font=("Arial", 14), command=cargar_datos)
    boton_cargar.pack(pady=20)

    root.mainloop()

iniciar_aplicacion()


